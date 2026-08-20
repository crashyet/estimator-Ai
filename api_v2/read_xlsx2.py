import sys, json
sys.path.insert(0, '/home/adhit/Desktop/Ngulik/magang_beecons/estimator/api_v2/.venv/lib/python3.14/site-packages')
import openpyxl
wb = openpyxl.load_workbook('/home/adhit/Desktop/Ngulik/magang_beecons/estimator/api_v2/ahsp/Item Pekerjaan CK.xlsx', data_only=True)
ws = wb.active
ids = set()
all_items = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is not None:
        id_val = str(row[0])
        name_val = str(row[1]) if row[1] else ''
        unit_val = str(row[2]) if row[2] else ''
        prefix = id_val.split('.')[0]
        ids.add(prefix)
        all_items.append((id_val, name_val, unit_val))
result = {"total": len(all_items), "prefixes": sorted(ids)}
samples = {}
for idx in [0,1,2,100,200,500,1000,1500,2000,2500,len(all_items)-3,len(all_items)-2,len(all_items)-1]:
    if idx < len(all_items):
        item = all_items[idx]
        samples[idx] = {"id": item[0], "name": item[1][:100], "unit": item[2]}
result["samples"] = samples
with open('/home/adhit/Desktop/Ngulik/magang_beecons/estimator/api_v2/ahsp_analysis.json', 'w') as f:
    json.dump(result, f, indent=2, ensure_ascii=False)
print("DONE")
