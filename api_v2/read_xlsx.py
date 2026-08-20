#!/usr/bin/env python3
import sys
sys.path.insert(0, '/home/adhit/Desktop/Ngulik/magang_beecons/estimator/api_v2/.venv/lib/python3.14/site-packages')
import openpyxl

wb = openpyxl.load_workbook('/home/adhit/Desktop/Ngulik/magang_beecons/estimator/api_v2/ahsp/Item Pekerjaan CK.xlsx', data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'Sheet: {sheet_name} | Rows: {ws.max_row} | Cols: {ws.max_column}')
    headers = [str(c.value) for c in ws[1]]
    print(f'Headers: {headers}')
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(40, ws.max_row), values_only=True), 2):
        vals = [str(c)[:80] if c is not None else '' for c in row]
        print(f'  R{i}: {vals}')
    print(f'  ... total data rows: {ws.max_row}')
